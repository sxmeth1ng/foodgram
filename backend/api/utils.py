from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def export_shopping_cart(ingredients: list):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Список покупок'
    headers = ['Ингредиент', 'Ед. изм.', 'Количество']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    row_num = 2
    for item in ingredients:
        ws.cell(row=row_num, column=1, value=item['Ингредиент'])
        ws.cell(row=row_num, column=2, value=item['Ед.изм'])
        ws.cell(row=row_num, column=3, value=item['Количество'])
        row_num += 1
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                raise ValueError(f'Не удалось распознать значение в ячейке '
                                 f'{cell.column}:{cell.row}')
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
